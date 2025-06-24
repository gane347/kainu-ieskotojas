import openpyxl
import numpy as np
from openpyxl.styles import PatternFill
import argparse
import sys

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import undetected_chromedriver as uc
import time
import random
import re
import pandas as pd


def get_pigu_price_refined(price_text):
    # Prices in pigu.lt are formatted in a weird way. I use regex to format them correctly
    match = re.search(r'(\d+)\s*(\d+)\s*(€|EUR)?', price_text)

    if match:
        whole_part = match.group(1)
        decimal_part = match.group(2)
        # Combine the parts with a dot for float conversion
        result_str = f"{whole_part}.{decimal_part}"
        try:
            result = float(result_str)
            print(f"Got price: '{result_str}'")
            return result
        except ValueError:
            print(f"Could not convert '{result_str}' to float.")
            return None
    else:
        print("Could not parse the new price from the text.")
        return None


def human_like_delay(min_seconds=1, max_seconds=3):
    time.sleep(random.uniform(min_seconds, max_seconds))


def setup_colab_driver():
    try:
        # Regular Selenium with WebDriver Manager
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--window-size=1920,1080')
        options.add_argument(
            '--user-agent='
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
            'AppleWebKit/537.36 (KHTML, like Gecko) '
            'Chrome/120.0.0.0 Safari/537.36')

        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=options)
        print("Using regular Selenium with WebDriver Manager")
        return driver

    except Exception as ex:
        print(f"WebDriver Manager failed: {ex}")

        try:
            # System Chrome
            options = Options()
            options.add_argument('--headless')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-gpu')
            options.add_argument('--window-size=1920,1080')
            options.binary_location = '/usr/bin/chromium-browser'

            driver = webdriver.Chrome(options=options)
            print("Using system Chrome")
            return driver

        except Exception as ex:
            print(f"System Chrome failed: {ex}")
            raise Exception("All Chrome driver methods failed. Please check installation.")


def setup_driver(headless=True):
    options = uc.ChromeOptions()
    if headless:
        options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--user-agent=Mozilla/5.0 '
                         '(Windows NT 10.0; Win64; x64) '
                         'AppleWebKit/537.36 (KHTML, like Gecko) '
                         'Chrome/120.0.0.0 Safari/537.36')
    driver = uc.Chrome(options=options)
    return driver


def get_prices(links, captcha):
    result = {}
    driver = setup_colab_driver()
    progress = 1

    try:
        for link in links:
            print(f"Scraping page: {link}\n")
            sys.stdout.write(f"Progress: {progress}/{len(links)+1}\n")
            sys.stdout.flush()
            progress += 1
            # Assign a default value for bad URLs
            if pd.isna(link) or link is None or (isinstance(link, str) and link.strip() == ""):
                print(f"--> Bad URL detected: '{link}'. Setting price to '0'.")
                result[link] = '0'
                continue

            driver.get(link)
            human_like_delay(1, 2)  # Wait for page to load

            # Handle Cloudflare and CAPTCHA challenges
            page_source = driver.page_source.lower()
            page_title = driver.title.lower()

            if "just a moment" in page_title or "checking your browser" in page_source:
                print("Cloudflare challenge detected. Trying to bypass...")
                # Quit current headless driver
                driver.quit()
                # Start new driver with GUI (non-headless)
                driver = setup_driver(headless=False)
                driver.get(link)
                print("Browser window opened, please solve the CAPTCHA manually.")
                if captcha:
                    input("Press Enter after solving CAPTCHA to continue...")

            # Get price from source code
            try:
                selector = "[data-qa='product_page_price']"  # Selector for pigu.lt
                if "varle" in link:
                    selector = ".price-value"  # Selector for varle.lt
                new_price_element = WebDriverWait(driver, 2).until(
                    ec.presence_of_element_located((By.CSS_SELECTOR, selector))
                )

                # Extract the text content
                price_text = new_price_element.text
                if "pigu" in link:
                    result[link] = get_pigu_price_refined(price_text)
                else:
                    result[link] = price_text

            except Exception as ex:
                print(f"Error finding or processing price: {ex}")
                return None
    finally:
        try:
            driver.quit()
            return result
        except Exception as ex:
            print(f"Error closing file: {ex}")
            pass


if __name__ == "__main__":
    # Read variables from line
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=str, default='LIST.xlsx')
    parser.add_argument("--destination", type=str, default='LIST1.xlsx')
    parser.add_argument("--from_index", type=int, default=2)
    parser.add_argument("--to_index", type=int, default=9999)
    parser.add_argument("--captchas", type=bool, default=False)
    parser.add_argument("--new_price", type=int, default=6)
    parser.add_argument("--old_price", type=int, default=5)
    parser.add_argument("--url", type=int, default=4)
    parser.add_argument("--compare", type=bool, default=True)
    parser.add_argument("--find", type=bool, default=True)
    args = parser.parse_args()

    # Define variables
    source_file_path = args.source
    destination_file_path = args.destination
    from_index = args.from_index
    to_index = args.to_index
    solve_captcha = args.captchas
    url_row = args.url
    write_price_row = args.new_price
    compare_price_row = args.old_price
    compare_price = args.compare
    find_new_prices = args.find

    try:
        # Read the Excel file
        workbook = openpyxl.load_workbook(source_file_path)
        sheet = workbook.active

        urls = []

        start_row = int(np.max([2, from_index]))
        end_row = int(np.min([sheet.max_row, to_index]))

        if find_new_prices:
            # Get urls
            for row_index in range(start_row, end_row + 1):
                url = sheet.cell(row=row_index, column=url_row).value
                if url:
                    urls.append(url)

            # Write column header
            if sheet.cell(row=1, column=write_price_row).value != 'New price':
                sheet.cell(row=1, column=write_price_row).value = 'New price'

            # Find price of products
            prices = get_prices(urls, solve_captcha)
            print(prices)

            # Write new price
            for row_index in range(start_row, end_row + 1):
                url = sheet.cell(row=row_index, column=url_row).value
                if url in prices:
                    sheet.cell(row=row_index, column=write_price_row).value = prices[url]

        if compare_price:
            # Set colors
            redFill = PatternFill(start_color='ff9696',
                                  end_color='ff9696',
                                  fill_type='solid')
            greenFill = PatternFill(start_color='9fff96',
                                    end_color='9fff96',
                                    fill_type='solid')

            # Scan each line and compare prices - if price has decreased color it green, if increased color it red
            for row_index in range(start_row, end_row + 1):
                old_value_text = sheet.cell(row=row_index, column=compare_price_row).value
                new_value_text = sheet.cell(row=row_index, column=write_price_row).value
                if old_value_text and new_value_text:
                    old_price = float(old_value_text)
                    new_price = float(new_value_text)
                    if old_price > new_price:
                        sheet.cell(row=row_index, column=write_price_row).fill = greenFill
                    elif old_price < new_price:
                        sheet.cell(row=row_index, column=write_price_row).fill = redFill

        # Save data
        workbook.save(destination_file_path)
        print(f"Successfully processed the file '{source_file_path}' by iterating.")
        sys.stdout.write("FINISHED\n")
        sys.stdout.flush()

    except FileNotFoundError:
        print(f"Error: The file '{source_file_path}' was not found.")
    except KeyError as e:
        print(f"Error: A column was not found. Please check your column indices. Details: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")
